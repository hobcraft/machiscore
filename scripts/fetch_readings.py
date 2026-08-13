"""
総務省「全国地方公共団体コード」から市区町村名の読み仮名を取得する。

かな・ローマ字での検索を可能にするために使う。e-Statのデータには読みが
含まれないため、こちらを出典にする。

出力: ../data/readings.json
  {"13113": {"kana": "シブヤク", "hiragana": "しぶやく", "romaji": "shibuyaku"}, ...}

注意:
  * 総務省の配布ファイルはコンテンツID（000925835.xlsx など）で、更新のたびに
    変わる。URLを直書きすると静かに古いデータを使い続けることになるので、
    一覧ページからリンクを辿り、ヘッダー行の構造で目的のファイルかを検証する。
  * 団体コードは6桁（末尾は検査数字）。先頭5桁がe-Stat側の市区町村コードにあたる。
"""
import json
import re
import unicodedata
from pathlib import Path

import openpyxl
import requests

INDEX_URL = "https://www.soumu.go.jp/denshijiti/code.html"
BASE_URL = "https://www.soumu.go.jp"

# このファイルだと判定するためのヘッダー（改行や全角を除いて比較する）
REQUIRED_HEADERS = ["団体コード", "都道府県名（漢字）", "市区町村名（漢字）", "市区町村名（カナ）"]

# 2021年の統計は浜松市の旧7区だが、コード表は2024年の再編後（3区）なので拾えない。
# 件数が少なく機械的に補えるため、ここで明示的に補完する。
MANUAL_READINGS = {
    "22131": "ハママツシナカク",
    "22132": "ハママツシヒガシク",
    "22133": "ハママツシニシク",
    "22134": "ハママツシミナミク",
    "22135": "ハママツシキタク",
    "22136": "ハママツシハマキタク",
    "22137": "ハママツシテンリュウク",
}

OUT_DIR = Path(__file__).resolve().parent.parent / "data"
OUT_DIR.mkdir(exist_ok=True)

# ヘボン式のローマ字変換表。長いものから順に当てる。
_ROMAJI = {
    "キャ": "kya", "キュ": "kyu", "キョ": "kyo", "シャ": "sha", "シュ": "shu", "ショ": "sho",
    "チャ": "cha", "チュ": "chu", "チョ": "cho", "ニャ": "nya", "ニュ": "nyu", "ニョ": "nyo",
    "ヒャ": "hya", "ヒュ": "hyu", "ヒョ": "hyo", "ミャ": "mya", "ミュ": "myu", "ミョ": "myo",
    "リャ": "rya", "リュ": "ryu", "リョ": "ryo", "ギャ": "gya", "ギュ": "gyu", "ギョ": "gyo",
    "ジャ": "ja", "ジュ": "ju", "ジョ": "jo", "ビャ": "bya", "ビュ": "byu", "ビョ": "byo",
    "ピャ": "pya", "ピュ": "pyu", "ピョ": "pyo", "ヂャ": "ja", "ヂュ": "ju", "ヂョ": "jo",
    "ア": "a", "イ": "i", "ウ": "u", "エ": "e", "オ": "o",
    "カ": "ka", "キ": "ki", "ク": "ku", "ケ": "ke", "コ": "ko",
    "サ": "sa", "シ": "shi", "ス": "su", "セ": "se", "ソ": "so",
    "タ": "ta", "チ": "chi", "ツ": "tsu", "テ": "te", "ト": "to",
    "ナ": "na", "ニ": "ni", "ヌ": "nu", "ネ": "ne", "ノ": "no",
    "ハ": "ha", "ヒ": "hi", "フ": "fu", "ヘ": "he", "ホ": "ho",
    "マ": "ma", "ミ": "mi", "ム": "mu", "メ": "me", "モ": "mo",
    "ヤ": "ya", "ユ": "yu", "ヨ": "yo",
    "ラ": "ra", "リ": "ri", "ル": "ru", "レ": "re", "ロ": "ro",
    "ワ": "wa", "ヲ": "o", "ン": "n",
    "ガ": "ga", "ギ": "gi", "グ": "gu", "ゲ": "ge", "ゴ": "go",
    "ザ": "za", "ジ": "ji", "ズ": "zu", "ゼ": "ze", "ゾ": "zo",
    "ダ": "da", "ヂ": "ji", "ヅ": "zu", "デ": "de", "ド": "do",
    "バ": "ba", "ビ": "bi", "ブ": "bu", "ベ": "be", "ボ": "bo",
    "パ": "pa", "ピ": "pi", "プ": "pu", "ペ": "pe", "ポ": "po",
    "ァ": "a", "ィ": "i", "ゥ": "u", "ェ": "e", "ォ": "o",
    "ャ": "ya", "ュ": "yu", "ョ": "yo", "ー": "",
}


def to_romaji(kana):
    """カタカナをヘボン式ローマ字にする。促音「ッ」は次の子音を重ねる。"""
    out = []
    i = 0
    while i < len(kana):
        if kana[i] == "ッ":
            # 次の音の頭の子音を重ねる（サッポロ -> sapporo）
            nxt = _lookup(kana, i + 1)
            if nxt and nxt[1] and nxt[1][0] not in "aiueo":
                out.append(nxt[1][0])
            i += 1
            continue
        hit = _lookup(kana, i)
        if hit:
            length, roman = hit
            out.append(roman)
            i += length
        else:
            i += 1  # 変換表にない文字は落とす
    return "".join(out)


def _lookup(kana, i):
    for length in (2, 1):
        chunk = kana[i : i + length]
        if chunk in _ROMAJI:
            return length, _ROMAJI[chunk]
    return None


def to_hiragana(kana):
    return "".join(
        chr(ord(c) - 0x60) if "ァ" <= c <= "ヶ" else c for c in kana
    )


def find_code_workbook():
    """一覧ページからコード表のExcelを探し、ヘッダー構造で正しさを確かめる。"""
    res = requests.get(INDEX_URL, timeout=60)
    res.raise_for_status()
    html = res.content.decode("shift_jis", errors="replace")
    links = re.findall(r'href="(/main_content/\d+\.xlsx)"', html)
    if not links:
        raise RuntimeError(f"{INDEX_URL} にExcelへのリンクが見つかりません")

    for link in links:
        content = requests.get(BASE_URL + link, timeout=120).content
        path = OUT_DIR / "_soumu_code.xlsx"
        path.write_bytes(content)
        workbook = openpyxl.load_workbook(path, read_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        header = next(sheet.iter_rows(values_only=True), ())
        flat = [re.sub(r"\s", "", str(c)) for c in header if c]
        if all(any(h == f for f in flat) for h in REQUIRED_HEADERS):
            print(f"コード表を特定: {link}")
            return workbook
        workbook.close()

    raise RuntimeError(
        "期待するヘッダーを持つExcelが見つかりません。"
        "総務省のページ構成かファイル形式が変わった可能性があります。"
    )


def main():
    workbook = find_code_workbook()
    readings = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            code, _pref, _city, _pref_kana, city_kana = (list(row) + [None] * 5)[:5]
            if not code or not city_kana:
                continue
            # 半角カタカナで配布されているので全角に正規化する
            readings[str(code)[:5]] = unicodedata.normalize("NFKC", str(city_kana))
    workbook.close()
    readings.update(MANUAL_READINGS)
    print(f"読み仮名: {len(readings)}件")

    records = {
        code: {
            "kana": kana,
            "hiragana": to_hiragana(kana),
            "romaji": to_romaji(kana),
        }
        for code, kana in readings.items()
    }

    out_path = OUT_DIR / "readings.json"
    out_path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"書き出し: {out_path}")
    for code in ("13113", "01101", "22137"):
        if code in records:
            print(f"  {code}: {records[code]}")


if __name__ == "__main__":
    main()
